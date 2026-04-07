import base64
import concurrent.futures
import os
import random
import re
import tempfile
import time
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any, Optional

import uvicorn
from curl_cffi import requests
from curl_cffi.requests.errors import RequestsError
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field


app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


PROXIES = {
    "http": "http://127.0.0.1:10808",
    "https": "http://127.0.0.1:10808",
}

DEFAULT_COOKIES = "_ga=GA1.1.741989139.1770273798; _bl_uid=a0mqIl359yk3F8ba29Rb6nI4RwRX; _tt_enable_cookie=1; _ttp=01KGP8N33PX9W91ZK9BMVT8BAD_.tt.1; _fbp=fb.1.1770273803492.904678892400498813; AGL_USER_ID=537f2997-a101-4a67-a4d6-df3d518fa256; appVersion=2.0; deviceType=pc; smidV2=20260205144652e4c9fce2e5abf9ac3f906db04d61c1630063f81d328030f00; deviceId=b34e10427af24b6b484efb90e070b316; _c_WBKFRo=rrYPw9CBxUTJ08Ez1u9d5cXqkcsTPKAU4pJLYYjC; _gcl_au=1.1.1470792634.1770273798.1354150457.1770274152.1770274152; page_session=ae2b0fc3-8ade-4934-b604-454ecc6308af; SESSION=YmVhNTNjZTgtNTM0NS00NDRhLWIwZjUtOTQ0YWU2YWE0ODY5; Hm_lvt_8aa1693861618ac63989ae373e684811=1775555364; HMACCOUNT=81DBD1F1490DA564; _twpid=tw.1775555364313.490781423161441852; _cfuvid=DYVFxrDiOsxIFLpVl6j4CFbJCy7YvdhskyAUhuh8ahc-1775555364.8915012-1.0.1.1-51ogg.P5Lwd4WbVe2QubtUqRk82GoD2VwL7bt3Cxqr0; _clck=11ni9t1%5E2%5Eg50%5E0%5E2227; cf_clearance=XJEJEdyU2cXM5AmhcJ.6XFZ8SFFAVvXIFShLK4dGWao-1775556833-1.2.1.1-KC4g7caKEEFZUgveqYejeiTrW7aKURou.oQChHM.vKTz6m.Ae9DU3qIObMOMc6AKpTt4mVM5giPomfJ18NYeKEfrZ0Fw27HkADfHsW.cQIfINaUPPIT2MbnFYt8YtX7v_m0gVXrUht7RXmrH6HEnyalyrxNVfqPQgk.7UTmzyB1K6vREWsSWzISd4UqTKZVC6qvwxBHNrEny9Ygv2rDrrlLJtUYXizYtmOzTf2npqv9BGLxm41hRDNLF_xUVVIfE4wrH8GMB4DVL0GoSgRnje0aaskI2W16mrkYVJDtQg9VukgTv.Mdh3imIISZMO9rHrc9mHFhQV_iqyRvyT9qmEw; ttcsid=1775555364919::it7L5JTGwc_07FO87jQN.3.1775556834392.0::1.1468053.1468933::1460772.17.0.0::0.0.0; ttcsid_CM9SHDBC77U4KJBR96OG=1775555364918::SBO5myw1k3_nB2C3XSGz.3.1775556834392.1; _clsk=23vq9d%5E1775556836692%5E3%5E1%5Ek.clarity.ms%2Fcollect; Hm_lpvt_8aa1693861618ac63989ae373e684811=1775556847; _ga_Q21FRKKG88=GS2.1.s1775555364$o5$g1$t1775556847$j40$l0$h0; _uetsid=0f7abe10326711f19b55e3fbf7b3a8a5; _uetvid=f69655e0025d11f18ba8fb211782ed3e; .thumbcache_211a882976e013454a0403b9c1967076=W+JRzQbNhpnW6nYGmbYuN8J+xCh/vT/RDj1XY+dprgtdI1Eqsh5b903XHHDlhQtBgLjsz5358Q4SKvli2wvvfA%3D%3D"
BLOCKED_STATUS_CODES = {401, 403, 429, 503}
BLOCKED_RESPONSE_MARKERS = (
    "just a moment",
    "attention required",
    "verify you are human",
    "captcha",
    "cf-chl",
    "cloudflare",
    "访问过于频繁",
    "访问受限",
)
EXPORT_PID_DELAY_RANGE = (1.2, 2.4)
EXPORT_REQUEST_RETRIES = 2
SCRAPE_DETAIL_MAX_WORKERS = 4
EXPORT_CACHE_DIR = os.path.join(os.path.dirname(__file__), "_export_cache")
os.makedirs(EXPORT_CACHE_DIR, exist_ok=True)


UNIT_MULTIPLIERS = {
    "\u4ebf": 100000000,
    "\u4e07": 10000,
    "\u5343": 1000,
    "k": 1000,
    "K": 1000,
    "w": 10000,
    "W": 10000,
    "m": 1000000,
    "M": 1000000,
}

CURRENCY_SYMBOLS = ("$", "\u00a5", "\uffe5", "\u20ac", "\u00a3")
EXPORT_RANGE_CONFIG = {
    "yesterday": {"label": "\u6628\u65e5", "days": 1},
    "recent_3": {"label": "\u8fd13\u5929", "days": 3},
    "recent_7": {"label": "\u8fc7\u53bb7\u5929", "days": 7},
    "recent_30": {"label": "\u8fc7\u53bb30\u5929", "days": 30},
    "all": {"label": "\u5168\u65f6\u95f4\u6bb5", "days": None},
}
COUNTRY_LABELS = {
    "US": "\u7f8e\u56fd",
    "MX": "\u58a8\u897f\u54e5",
    "GB": "\u82f1\u56fd",
    "ES": "\u897f\u73ed\u7259",
    "DE": "\u5fb7\u56fd",
    "FR": "\u6cd5\u56fd",
    "IT": "\u610f\u5927\u5229",
}


class ScrapeRequest(BaseModel):
    pid: str
    startDate: str
    endDate: str
    sortBy: str
    pageNo: int = 1
    pageSize: int = 10
    cookie: Optional[str] = None
    country: str = "US"
    currency: str = "USD"


class ExportRequest(BaseModel):
    pids: list[str] = Field(default_factory=list)
    timeRange: str = "recent_7"
    cookie: Optional[str] = None
    country: str = "US"
    currency: str = "USD"


class AntiBotError(RuntimeError):
    pass


def get_headers(custom_cookie: Optional[str], pid: str, country: str = "US", currency: str = "USD") -> dict[str, str]:
    cookie = custom_cookie.strip() if custom_cookie and custom_cookie.strip() else DEFAULT_COOKIES
    return {
        "accept": "application/json, text/plain, */*",
        "accept-language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
        "content-type": "application/json",
        "country": country,
        "currency": currency,
        "language": "zh-CN",
        "origin": "https://www.kalodata.com",
        "referer": f"https://www.kalodata.com/product/detail?id={pid}&language=zh-CN&currency={currency}&region={country}",
        "sec-ch-ua": '"Not:A-Brand";v="99", "Microsoft Edge";v="145", "Chromium";v="145"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-origin",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36 Edg/145.0.0.0",
        "cookie": cookie,
    }


def parse_compact_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0

    for symbol in CURRENCY_SYMBOLS:
        text = text.replace(symbol, "")
    text = text.strip()

    match = re.match(r"^([-+]?\d+(?:\.\d+)?)([\u4ebf\u4e07\u5343kKwWmM]?)$", text)
    if match:
        number = float(match.group(1))
        unit = match.group(2)
        return number * UNIT_MULTIPLIERS.get(unit, 1)

    cleaned = re.sub(r"[^0-9.\-+]", "", text)
    if not cleaned:
        return 0.0
    return float(cleaned)


def safe_display(value: Any, fallback: str = "\u672a\u77e5") -> str:
    if value is None:
        return fallback
    text = str(value).strip()
    return text if text else fallback


def normalize_int(value: Any) -> int:
    return int(round(parse_compact_number(value)))


def get_primary_category(category: Any) -> str:
    text = safe_display(category, "\u672a\u77e5")
    if text == "\u672a\u77e5":
        return text
    return text.split(" > ")[0].strip() or text


def normalize_handle(value: Any) -> str:
    text = safe_display(value, "").strip()
    return text.lstrip("@")


def extract_handle(data: Any) -> str:
    if not isinstance(data, dict):
        return ""
    for key in ("handle", "author_handle", "authorHandle", "unique_id", "authorUniqueId", "username"):
        handle = normalize_handle(data.get(key))
        if handle:
            return handle
    return ""


def build_tiktok_video_link(handle: str, video_id: Any) -> str:
    clean_handle = normalize_handle(handle)
    clean_video_id = safe_display(video_id, "")
    if not clean_handle or not clean_video_id:
        return ""
    return f"https://www.tiktok.com/@{clean_handle}/video/{clean_video_id}"


def calc_growth(cur: float, prev: float) -> float:
    if prev == 0:
        return 100.0 if cur > 0 else 0.0
    return ((cur - prev) / prev) * 100.0


def get_data_base_time(now: Optional[datetime] = None) -> datetime:
    return (now or datetime.now()) - timedelta(days=2)


def get_dt_str(days_ago: int, now: Optional[datetime] = None) -> str:
    return ((now or datetime.now()) - timedelta(days=days_ago)).strftime("%Y-%m-%d")


def build_comparison_windows(now: Optional[datetime] = None) -> dict[str, tuple[str, str]]:
    anchor = now or datetime.now()
    return {
        "cur_3": (get_dt_str(4, anchor), get_dt_str(2, anchor)),
        "prev_3": (get_dt_str(7, anchor), get_dt_str(5, anchor)),
        "cur_7": (get_dt_str(8, anchor), get_dt_str(2, anchor)),
        "prev_7": (get_dt_str(15, anchor), get_dt_str(9, anchor)),
    }


def build_export_range(time_range: str) -> tuple[str, str, str]:
    config = EXPORT_RANGE_CONFIG.get(time_range, EXPORT_RANGE_CONFIG["recent_7"])
    end_dt = get_data_base_time()
    if config["days"] is None:
        start_dt = datetime(2000, 1, 1)
    else:
        start_dt = end_dt - timedelta(days=config["days"] - 1)
    return start_dt.strftime("%Y-%m-%d"), end_dt.strftime("%Y-%m-%d"), config["label"]


def is_meaningful_info(info_data: dict[str, Any]) -> bool:
    if not info_data:
        return False
    candidate_keys = (
        "product_title",
        "brand_name",
        "unit_price",
        "collect_day",
        "pri_cate_id",
        "sec_cate_id",
        "ter_cate_id",
        "shop_name",
        "store_name",
    )
    return any(info_data.get(key) for key in candidate_keys)


def build_retry_delay(attempt: int) -> float:
    return min(8.0, 1.2 * (2**attempt)) + random.uniform(0.2, 0.8)


def looks_like_block_page(text: str) -> bool:
    lowered = text.lower()
    return any(marker in lowered for marker in BLOCKED_RESPONSE_MARKERS)


def looks_like_block_payload(payload: Any) -> bool:
    if not isinstance(payload, dict):
        return False

    status = payload.get("status")
    code = payload.get("code")
    if status in BLOCKED_STATUS_CODES or code in BLOCKED_STATUS_CODES:
        return True

    values_to_check = [
        payload.get("message"),
        payload.get("error"),
        payload.get("detail"),
    ]
    return any(looks_like_block_page(str(value)) for value in values_to_check if value)


def request_json(
    method: str,
    url: str,
    headers: dict[str, str],
    payload: Optional[dict[str, Any]] = None,
    timeout: int = 10,
    retries: int = 0,
) -> Any:
    last_error: Optional[Exception] = None

    for attempt in range(retries + 1):
        if attempt:
            time.sleep(build_retry_delay(attempt - 1))

        try:
            if method == "POST":
                response = requests.post(
                    url,
                    json=payload,
                    headers=headers,
                    proxies=PROXIES,
                    timeout=timeout,
                    impersonate="edge101",
                )
            else:
                response = requests.get(
                    url,
                    headers=headers,
                    proxies=PROXIES,
                    timeout=timeout,
                    impersonate="edge101",
                )

            status_code = getattr(response, "status_code", 200)
            response_text = getattr(response, "text", "") or ""
            if status_code in BLOCKED_STATUS_CODES or looks_like_block_page(response_text):
                raise AntiBotError(f"Request blocked: {method} {url} -> HTTP {status_code}")

            data = response.json()
            if looks_like_block_payload(data):
                raise AntiBotError(f"Request blocked by payload: {method} {url}")
            return data
        except AntiBotError as exc:
            last_error = exc
            if attempt >= retries:
                raise
        except RequestsError as exc:
            last_error = exc
            if attempt >= retries:
                raise
        except Exception:
            raise

    if last_error:
        raise last_error
    raise RuntimeError(f"Request failed: {method} {url}")


def fetch_json(
    url: str,
    headers: dict[str, str],
    is_post: bool = True,
    payload: Optional[dict[str, Any]] = None,
    timeout: int = 10,
    retries: int = 0,
) -> Any:
    return request_json(
        "POST" if is_post else "GET",
        url,
        headers,
        payload=payload,
        timeout=timeout,
        retries=retries,
    )


def default_total_metrics() -> dict[str, Any]:
    return {
        "sale": 0.0,
        "revenue": 0.0,
        "sale_str": "0",
        "revenue_str": "$0",
        "day_sale_str": "0",
        "day_revenue_str": "$0",
    }


def fetch_total(
    pid: str,
    start_date: str,
    end_date: str,
    headers: dict[str, str],
    strict: bool = False,
) -> dict[str, Any]:
    payload = {
        "id": pid,
        "startDate": start_date,
        "endDate": end_date,
        "authority": True,
    }
    try:
        resp = fetch_json(
            "https://www.kalodata.com/product/detail/total",
            headers,
            True,
            payload,
            timeout=10,
            retries=EXPORT_REQUEST_RETRIES,
        )
        data = resp.get("data") or {}
        return {
            "sale": parse_compact_number(data.get("sale", 0)),
            "revenue": parse_compact_number(data.get("original_revenue", data.get("revenue", 0))),
            "sale_str": safe_display(data.get("sale"), "0"),
            "revenue_str": safe_display(data.get("revenue"), "$0"),
            "day_sale_str": safe_display(data.get("day_sale"), "0"),
            "day_revenue_str": safe_display(data.get("day_revenue"), "$0"),
        }
    except (AntiBotError, RequestsError):
        if strict:
            raise
    except Exception:
        if strict:
            raise
    return default_total_metrics()


def build_product_summary(
    pid: str,
    start_date: str,
    end_date: str,
    country: str,
    currency: str,
    cookie: Optional[str],
    sort_by: str = "revenue",
    safe_mode: bool = False,
) -> tuple[dict[str, Any], dict[str, str], dict[str, Any], bool]:
    headers = get_headers(cookie, pid, country, currency)
    count_payload = {
        "id": pid,
        "startDate": start_date,
        "endDate": end_date,
        "authority": True,
        "pageNo": 1,
        "pageSize": 10,
        "sort": [{"field": sort_by, "type": "DESC"}],
    }
    windows = build_comparison_windows()

    count_url = "https://www.kalodata.com/product/detail/video/count"
    creator_count_url = "https://www.kalodata.com/product/detail/creator/count"
    images_url = f"https://www.kalodata.com/product/detail/getImages?productId={pid}"
    detail_url = "https://www.kalodata.com/product/detail"

    def fetch_count_json() -> dict[str, Any]:
        return fetch_json(
            count_url,
            headers,
            True,
            count_payload,
            timeout=12,
            retries=EXPORT_REQUEST_RETRIES,
        )

    def fetch_creator_json() -> dict[str, Any]:
        return fetch_json(
            creator_count_url,
            headers,
            True,
            count_payload,
            timeout=10,
            retries=EXPORT_REQUEST_RETRIES,
        )

    def fetch_images_json() -> dict[str, Any]:
        return fetch_json(
            images_url,
            headers,
            False,
            None,
            timeout=10,
            retries=EXPORT_REQUEST_RETRIES,
        )

    def fetch_info_json() -> dict[str, Any]:
        return fetch_json(
            detail_url,
            headers,
            True,
            count_payload,
            timeout=12,
            retries=EXPORT_REQUEST_RETRIES,
        )

    try:
        if safe_mode:
            count_json = fetch_count_json()
            creator_json = fetch_creator_json()
            images_json = fetch_images_json()
            info_json = fetch_info_json()
            cur_range_data = fetch_total(pid, start_date, end_date, headers, strict=True)
            cur_3_data = fetch_total(pid, *windows["cur_3"], headers, strict=True)
            prev_3_data = fetch_total(pid, *windows["prev_3"], headers, strict=True)
            cur_7_data = fetch_total(pid, *windows["cur_7"], headers, strict=True)
            prev_7_data = fetch_total(pid, *windows["prev_7"], headers, strict=True)
        else:
            with concurrent.futures.ThreadPoolExecutor(max_workers=9) as executor:
                future_count = executor.submit(fetch_count_json)
                future_creator = executor.submit(fetch_creator_json)
                future_images = executor.submit(fetch_images_json)
                future_info = executor.submit(fetch_info_json)
                future_range = executor.submit(fetch_total, pid, start_date, end_date, headers)
                future_cur_3 = executor.submit(fetch_total, pid, *windows["cur_3"], headers)
                future_prev_3 = executor.submit(fetch_total, pid, *windows["prev_3"], headers)
                future_cur_7 = executor.submit(fetch_total, pid, *windows["cur_7"], headers)
                future_prev_7 = executor.submit(fetch_total, pid, *windows["prev_7"], headers)

                count_json = future_count.result()
                creator_json = future_creator.result()
                images_json = future_images.result()
                info_json = future_info.result()
                cur_range_data = future_range.result()
                cur_3_data = future_cur_3.result()
                prev_3_data = future_prev_3.result()
                cur_7_data = future_cur_7.result()
                prev_7_data = future_prev_7.result()
    except AntiBotError as exc:
        raise PermissionError("Kalodata blocked the request or the Cookie is expired.") from exc

    total_videos = count_json.get("data", 0) or 0
    total_creators = creator_json.get("data", 0) if isinstance(creator_json, dict) else 0

    images_data = images_json.get("data", []) if isinstance(images_json, dict) else []
    if isinstance(images_data, list):
        product_images = [str(item) for item in images_data if item]
    elif images_data:
        product_images = [str(images_data)]
    else:
        product_images = []

    info_data = info_json.get("data", {}) if isinstance(info_json, dict) else {}
    has_detail = is_meaningful_info(info_data)

    categories = [info_data.get("pri_cate_id", ""), info_data.get("sec_cate_id", ""), info_data.get("ter_cate_id", "")]
    product_category = " > ".join([item for item in categories if item]) or "\u672a\u77e5"
    product_title = safe_display(info_data.get("product_title"), "\u6682\u65e0\u7b80\u4ecb")
    store_name = safe_display(
        info_data.get("shop_name")
        or info_data.get("store_name")
        or info_data.get("seller_name")
        or info_data.get("brand_name"),
        "\u672a\u77e5",
    )
    collect_day = safe_display(info_data.get("collect_day"), "\u672a\u77e5")
    sku_info = info_data.get("skuInfo", [])
    stock = "\u672a\u77e5"
    if isinstance(sku_info, list) and sku_info:
        stock = safe_display(sku_info[0].get("stock"), "\u672a\u77e5")

    product = {
        "images": product_images,
        "totalCreators": total_creators,
        "totalVideos": total_videos,
        "price": safe_display(info_data.get("unit_price"), "\u672a\u77e5"),
        "category": product_category,
        "rating": safe_display(info_data.get("product_rating"), "\u6682\u65e0\u8bc4\u5206"),
        "title": product_title,
        "brand": store_name,
        "collectDay": collect_day,
        "stock": stock,
        "range_sale": cur_range_data["sale_str"],
        "range_revenue": cur_range_data["revenue_str"],
        "range_day_sale": cur_range_data["day_sale_str"],
        "range_day_revenue": cur_range_data["day_revenue_str"],
        "range_sale_value": cur_range_data["sale"],
        "range_revenue_value": cur_range_data["revenue"],
        "recent_3_sale": cur_3_data["sale_str"],
        "recent_3_revenue": cur_3_data["revenue_str"],
        "prev_3_sale": prev_3_data["sale_str"],
        "prev_3_revenue": prev_3_data["revenue_str"],
        "growth_3_sale": calc_growth(cur_3_data["sale"], prev_3_data["sale"]),
        "growth_3_revenue": calc_growth(cur_3_data["revenue"], prev_3_data["revenue"]),
        "recent_7_sale": cur_7_data["sale_str"],
        "recent_7_revenue": cur_7_data["revenue_str"],
        "prev_7_sale": prev_7_data["sale_str"],
        "prev_7_revenue": prev_7_data["revenue_str"],
        "growth_7_sale": calc_growth(cur_7_data["sale"], prev_7_data["sale"]),
        "growth_7_revenue": calc_growth(cur_7_data["revenue"], prev_7_data["revenue"]),
    }
    return product, headers, count_payload, has_detail


def fetch_video_detail_by_range(
    item: dict[str, Any],
    start_date: str,
    end_date: str,
    headers: dict[str, str],
) -> dict[str, Any]:
    video_id = item.get("id")
    mp4_url = "\u83b7\u53d6\u5931\u8d25"
    handle = "\u672a\u77e5"
    duration = "\u672a\u77e5"

    def get_mp4() -> str:
        resp = fetch_json(
            f"https://www.kalodata.com/video/detail/getVideoUrl?videoId={video_id}",
            headers,
            False,
            None,
            timeout=5,
            retries=1,
        )
        return resp.get("data", {}).get("url", "\u83b7\u53d6\u5931\u8d25")

    def get_detail() -> dict[str, Any]:
        resp = fetch_json(
            "https://www.kalodata.com/video/detail",
            headers,
            True,
            {
                "id": video_id,
                "startDate": start_date,
                "endDate": end_date,
                "authority": True,
            },
            timeout=5,
            retries=1,
        )
        return resp.get("data", {})

    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        future_mp4 = executor.submit(get_mp4)
        future_detail = executor.submit(get_detail)
        try:
            mp4_url = future_mp4.result()
        except Exception:
            pass
        try:
            detail_data = future_detail.result()
            handle = extract_handle(detail_data) or "\u672a\u77e5"
            duration = safe_display(detail_data.get("duration"), "\u672a\u77e5")
        except Exception:
            pass

    return {
        **item,
        "mp4Url": mp4_url,
        "handle": handle,
        "duration": duration,
        "tiktokVideoUrl": build_tiktok_video_link(handle, video_id) if handle != "\u672a\u77e5" else "\u672a\u77e5",
        "tiktokHomepageUrl": f"https://www.tiktok.com/@{handle}" if handle != "\u672a\u77e5" else "\u672a\u77e5",
        "coverImageUrl": f"https://img.kalocdn.com/tiktok.video/{video_id}/cover.png",
        "isAd": str(item.get("ad")) == "1",
    }


def fetch_video_detail(item: dict[str, Any], req: ScrapeRequest, headers: dict[str, str]) -> dict[str, Any]:
    return fetch_video_detail_by_range(item, req.startDate, req.endDate, headers)


def fetch_top_video_official_link(
    pid: str,
    start_date: str,
    end_date: str,
    headers: dict[str, str],
    count_payload: dict[str, Any],
) -> str:
    list_payload = {**count_payload, "pageNo": 1, "pageSize": 1}
    try:
        list_data = fetch_json(
            "https://www.kalodata.com/product/detail/video/queryList",
            headers,
            True,
            list_payload,
            timeout=15,
            retries=1,
        ).get("data", [])
        if not list_data:
            return ""

        first_video = list_data[0]
        video_id = first_video.get("id")
        if not video_id:
            return ""

        detailed_video = fetch_video_detail_by_range(first_video, start_date, end_date, headers)
        link = safe_display(detailed_video.get("tiktokVideoUrl"), "")
        return "" if link == "\u672a\u77e5" else link
    except Exception:
        return ""


def get_export_headers(time_label: str) -> list[str]:
    period_label = time_label
    return [
        "\u4ea7\u54c1PID",
        "\u56fd\u5bb6",
        "\u4ea7\u54c1\u4ef7\u683c",
        "\u5e97\u94fa\u540d\u79f0",
        "\u4ea7\u54c1\u7c7b\u76ee",
        "\u4ea7\u54c1\u8bc4\u5206",
        "\u6700\u65e9\u6536\u5f55\u65f6\u95f4",
        f"{period_label}\u89c6\u9891\u603b\u6570",
        f"{period_label}\u9500\u552e\u91cf",
        f"{period_label}\u9500\u552e\u989d",
        "\u8fd13\u5929\u8d70\u52bf",
        "\u8fd17\u5929\u8d70\u52bf",
        "\u8fd1\u671f\u7206\u6b3e\u89c6\u9891\u94fe\u63a5",
        "TikTok Shop\u94fe\u63a5",
        "\u4e3b\u56fe\u94fe\u63a5",
        "Tabcut\u94fe\u63a5",
    ]


def get_export_row_values(row: dict[str, Any]) -> list[Any]:
    return [
        row["pid"],
        row["country"],
        row["price"],
        row["brand"],
        row["category"],
        row["rating"],
        row["collectDay"],
        row["video_count"],
        row["range_sale"],
        row["range_revenue"],
        row["trend_3"],
        row["trend_7"],
        row["top_video_link"],
        row["tiktok_shop_link"],
        row["image_link"],
        row["tabcut_link"],
    ]


class IncrementalExportWorkbook:
    def __init__(self, time_label: str):
        temp_file = tempfile.NamedTemporaryFile(
            prefix="kalodata_export_",
            suffix=".xlsx",
            delete=False,
            dir=EXPORT_CACHE_DIR,
        )
        self.file_path = temp_file.name
        temp_file.close()

        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "\u6279\u91cf\u5bfc\u51fa"
        self.headers = get_export_headers(time_label)
        self.header_fill = PatternFill("solid", fgColor="1A7F82")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.body_alignment = Alignment(vertical="center", wrap_text=True)
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.column_widths: list[int] = []
        self.row_count = 0

        self.sheet.append(self.headers)
        for cell in self.sheet[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment

        self.sheet.freeze_panes = "A2"
        self.sheet.row_dimensions[1].height = 24
        for index, header in enumerate(self.headers, start=1):
            width = min(max(len(str(header)) + 4, 14), 36)
            self.column_widths.append(width)
            self.sheet.column_dimensions[get_column_letter(index)].width = width

        self._save()

    def _save(self) -> None:
        self.workbook.save(self.file_path)

    def append_row(self, row: dict[str, Any]) -> None:
        values = get_export_row_values(row)
        self.sheet.append(values)
        self.row_count += 1
        row_index = self.sheet.max_row

        for cell in self.sheet[row_index]:
            cell.alignment = self.body_alignment

        for index, value in enumerate(values, start=1):
            width = min(max(len(str(value or "")) + 4, 14), 36)
            if width > self.column_widths[index - 1]:
                self.column_widths[index - 1] = width
                self.sheet.column_dimensions[get_column_letter(index)].width = width

        self._save()

    def to_base64(self) -> str:
        self._save()
        with open(self.file_path, "rb") as workbook_file:
            return base64.b64encode(workbook_file.read()).decode("utf-8")

    def cleanup(self) -> None:
        try:
            os.remove(self.file_path)
        except OSError:
            pass


def build_export_workbook(rows: list[dict[str, Any]], time_label: str) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "\u6279\u91cf\u5bfc\u51fa"
    headers = get_export_headers(time_label)
    sheet.append(headers)

    for row in rows:
        sheet.append(get_export_row_values(row))

    header_fill = PatternFill("solid", fgColor="1A7F82")
    header_font = Font(color="FFFFFF", bold=True)
    body_alignment = Alignment(vertical="center", wrap_text=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 24

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = body_alignment

    for idx, column_cells in enumerate(sheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value_length = len(str(cell.value or ""))
            max_length = max(max_length, value_length)
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(max_length + 4, 14), 36)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return base64.b64encode(buffer.read()).decode("utf-8")


def build_export_row(pid: str, country: str, product: dict[str, Any], top_video_link: str) -> dict[str, Any]:
    image_link = ""
    if isinstance(product.get("images"), list) and product["images"]:
        image_link = str(product["images"][0])

    return {
        "pid": pid,
        "country": safe_display(country, "\u672a\u77e5"),
        "price": safe_display(product.get("price"), "\u672a\u77e5"),
        "brand": safe_display(product.get("brand"), "\u672a\u77e5"),
        "category": get_primary_category(product.get("category")),
        "rating": safe_display(product.get("rating"), "\u6682\u65e0"),
        "collectDay": safe_display(product.get("collectDay"), "\u672a\u77e5"),
        "video_count": normalize_int(product.get("totalVideos", 0)),
        "range_sale": normalize_int(product.get("range_sale_value", product.get("range_sale", 0))),
        "range_revenue": normalize_int(product.get("range_revenue_value", product.get("range_revenue", 0))),
        "trend_3": f"{float(product.get('growth_3_sale', 0.0)):+.1f}%",
        "trend_7": f"{float(product.get('growth_7_sale', 0.0)):+.1f}%",
        "top_video_link": top_video_link,
        "tiktok_shop_link": f"https://www.tiktok.com/shop/pdp/{pid}",
        "image_link": image_link,
        "tabcut_link": f"https://www.tabcut.com/zh-CN/ranking/goods/detail?id={pid}",
    }


@app.post("/api/scrape")
def scrape_data(req: ScrapeRequest):
    try:
        product, headers, count_payload, _ = build_product_summary(
            pid=req.pid,
            start_date=req.startDate,
            end_date=req.endDate,
            country=req.country,
            currency=req.currency,
            cookie=req.cookie,
            sort_by=req.sortBy,
        )

        total_videos = int(product.get("totalVideos", 0) or 0)
        if total_videos == 0:
            return {"total": 0, "list": [], "product": product}

        list_payload = {**count_payload, "pageNo": req.pageNo, "pageSize": req.pageSize}
        video_list = fetch_json(
            "https://www.kalodata.com/product/detail/video/queryList",
            headers,
            True,
            list_payload,
            timeout=15,
            retries=1,
        ).get("data", [])

        worker_count = min(SCRAPE_DETAIL_MAX_WORKERS, len(video_list)) or 1
        with concurrent.futures.ThreadPoolExecutor(max_workers=worker_count) as executor:
            detailed_list = list(executor.map(lambda item: fetch_video_detail(item, req, headers), video_list))

        return {
            "total": total_videos,
            "list": detailed_list,
            "product": product,
        }
    except (PermissionError, AntiBotError):
        return JSONResponse(
            status_code=403,
            content={"error": "Kalodata \u8bbf\u95ee\u88ab\u62e6\u622a\u6216 Cookie \u5df2\u8fc7\u671f\uff0c\u8bf7\u5728\u8bbe\u7f6e\u4e2d\u66f4\u65b0\u6709\u6548 Cookie\u3002"},
        )
    except RequestsError as exc:
        return JSONResponse(
            status_code=500,
            content={"error": f"\u7f51\u7edc\u5f02\u5e38\u6216\u4ee3\u7406\u8fde\u63a5\u5931\u8d25\uff0c\u8bf7\u786e\u8ba4 10808 \u7aef\u53e3\u4ee3\u7406\u662f\u5426\u5f00\u542f: {str(exc)}"},
        )
    except Exception as exc:
        return JSONResponse(status_code=500, content={"error": f"\u6293\u53d6\u5931\u8d25: {str(exc)}"})


@app.post("/api/export")
def export_products(req: ExportRequest):
    pids = []
    seen = set()
    for raw_pid in req.pids:
        pid = str(raw_pid).strip()
        if pid and pid not in seen:
            seen.add(pid)
            pids.append(pid)

    if not pids:
        return JSONResponse(status_code=400, content={"error": "\u8bf7\u81f3\u5c11\u8f93\u5165\u4e00\u4e2a PID\u3002"})

    start_date, end_date, time_label = build_export_range(req.timeRange)
    failed_pids: list[str] = []
    success_count = 0
    stopped_early = False
    warning_message = None
    export_book: Optional[IncrementalExportWorkbook] = None

    def process_pid(pid: str) -> Optional[dict[str, Any]]:
        try:
            product, headers, count_payload, has_detail = build_product_summary(
                pid=pid,
                start_date=start_date,
                end_date=end_date,
                country=req.country,
                currency=req.currency,
                cookie=req.cookie,
                sort_by="revenue",
                safe_mode=True,
            )
            if not has_detail:
                return None
            top_video_link = fetch_top_video_official_link(
                pid=pid,
                start_date=start_date,
                end_date=end_date,
                headers=headers,
                count_payload=count_payload,
            )
            return build_export_row(pid, req.country, product, top_video_link)
        except PermissionError:
            raise
        except Exception:
            return None

    try:
        for index, pid in enumerate(pids):
            try:
                row = process_pid(pid)
            except PermissionError:
                failed_pids.append(pid)
                failed_pids.extend(pids[index + 1 :])
                stopped_early = True
                warning_message = "\u68c0\u6d4b\u5230 Kalodata \u53cd\u722c\u6216 Cookie \u5df2\u5931\u6548\uff0c\u5df2\u63d0\u524d\u505c\u6b62\u672c\u6b21\u6279\u91cf\u5bfc\u51fa\uff0c\u524d\u9762\u5df2\u6293\u5230\u7684\u6570\u636e\u5df2\u4fdd\u7559\u3002"
                break

            if row is None:
                failed_pids.append(pid)
            else:
                if export_book is None:
                    export_book = IncrementalExportWorkbook(time_label)
                export_book.append_row(row)
                success_count += 1

            if index < len(pids) - 1:
                time.sleep(random.uniform(*EXPORT_PID_DELAY_RANGE))
    finally:
        file_name = None
        file_content_base64 = None
        if export_book and export_book.row_count > 0:
            try:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = f"\u6279\u91cf\u5bfc\u51fa_{time_label}_{timestamp}.xlsx"
                file_content_base64 = export_book.to_base64()
            finally:
                export_book.cleanup()

    return {
        "timeLabel": time_label,
        "successCount": success_count,
        "failureCount": len(failed_pids),
        "failedPids": failed_pids,
        "fileName": file_name,
        "fileContentBase64": file_content_base64,
        "stoppedEarly": stopped_early,
        "warning": warning_message,
    }


dist_dir = os.path.join(os.path.dirname(__file__), "dist")
assets_path = os.path.join(dist_dir, "assets")
if os.path.exists(assets_path):
    app.mount("/assets", StaticFiles(directory=assets_path), name="assets")


@app.get("/{catchall:path}")
def serve_spa(catchall: str):
    if catchall.startswith("api/"):
        return JSONResponse({"detail": "API endpoint not found"}, status_code=404)

    file_path = os.path.join(dist_dir, catchall)
    if os.path.isfile(file_path):
        return FileResponse(file_path)

    index_file = os.path.join(dist_dir, "index.html")
    if os.path.exists(index_file):
        return FileResponse(index_file)

    return JSONResponse(
        {"error": "\u672a\u627e\u5230\u524d\u7aef\u6784\u5efa\u4ea7\u7269\uff0c\u8bf7\u5148\u5728\u524d\u7aef\u76ee\u5f55\u6267\u884c npm run build\u3002"},
        status_code=404,
    )


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8010)
